"""
Output generation utilities for Tableau metadata.

Supports JSON, Excel, and HTML output formats.
"""

import json
from typing import Optional, Dict, Any, List
from pathlib import Path
from datetime import datetime

from models.metadata_models import WorkbookMetadata


class OutputGenerator:
    """
    Generates various output formats for extracted metadata.
    
    Supported formats:
    - JSON (detailed, machine-readable)
    - Excel (summary workbook with multiple sheets)
    - HTML (interactive report)
    - Text (console summary)
    """
    
    def __init__(self, metadata: WorkbookMetadata):
        """
        Initialize the output generator.
        
        Args:
            metadata: Extracted workbook metadata
        """
        self.metadata = metadata
    
    def to_json(self, output_path: Optional[str] = None, indent: int = 2) -> str:
        """
        Export metadata to JSON format.
        
        Args:
            output_path: Optional path to save the file
            indent: JSON indentation level
            
        Returns:
            str: JSON string
        """
        json_str = self.metadata.model_dump_json(indent=indent)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
        
        return json_str
    
    def to_dict(self) -> Dict[str, Any]:
        """Export metadata to dictionary."""
        return self.metadata.model_dump()
    
    def to_excel(self, output_path: str):
        """
        Export metadata to Excel workbook with multiple sheets.
        
        Args:
            output_path: Path to save the Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, header_font, header_fill)
        
        # Fields sheet
        ws_fields = wb.create_sheet("Fields")
        self._write_fields_sheet(ws_fields, header_font, header_fill)
        
        # Calculated Fields sheet
        ws_calcs = wb.create_sheet("Calculated Fields")
        self._write_calculated_fields_sheet(ws_calcs, header_font, header_fill)
        
        # Sheets sheet
        ws_sheets = wb.create_sheet("Worksheets")
        self._write_worksheets_sheet(ws_sheets, header_font, header_fill)
        
        # Filters sheet
        ws_filters = wb.create_sheet("Filters")
        self._write_filters_sheet(ws_filters, header_font, header_fill)
        
        # Dashboards sheet
        ws_dashboards = wb.create_sheet("Dashboards")
        self._write_dashboards_sheet(ws_dashboards, header_font, header_fill)
        
        # Parameters sheet
        ws_params = wb.create_sheet("Parameters")
        self._write_parameters_sheet(ws_params, header_font, header_fill)
        
        # Relationships sheet
        ws_rels = wb.create_sheet("Relationships")
        self._write_relationships_sheet(ws_rels, header_font, header_fill)
        
        # Metrics sheet (one row per metric-worksheet combination)
        ws_metrics = wb.create_sheet("Metrics")
        self._write_metrics_sheet(ws_metrics, header_font, header_fill)
        
        # KPI Summary sheet (business-friendly overview of all KPIs/metrics with calculations)
        ws_kpi_summary = wb.create_sheet("KPI Summary")
        self._write_kpi_summary_sheet(ws_kpi_summary, header_font, header_fill)
        
        wb.save(output_path)
    
    def _write_summary_sheet(self, ws, header_font, header_fill):
        """Write summary information."""
        from openpyxl.styles import Font
        
        data = [
            ["Workbook Name", self.metadata.name],
            ["Version", self.metadata.version or "N/A"],
            ["Source File", self.metadata.source_file or "N/A"],
            ["Extraction Time", str(self.metadata.extraction_timestamp) if self.metadata.extraction_timestamp else "N/A"],
            ["Extraction Method", self.metadata.extraction_method],
            [""],
            ["Statistics", ""],
            ["Total Data Sources", len(self.metadata.datasources)],
            ["Total Sheets", self.metadata.total_sheets],
            ["Total Dashboards", self.metadata.total_dashboards],
            ["Total Fields", self.metadata.total_fields],
            ["Total Calculated Fields", self.metadata.total_calculated_fields],
            ["Total Parameters", self.metadata.total_parameters],
            ["Total Filters", self.metadata.total_filters],
        ]
        
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:
                    cell.font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _write_fields_sheet(self, ws, header_font, header_fill):
        """Write fields information."""
        headers = ["Data Source", "Field Name", "Caption", "Data Type", "Role", "Default Aggregation", "Hidden"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for ds in self.metadata.datasources:
            for field in ds.fields:
                ws.cell(row=row_idx, column=1, value=ds.display_name)
                ws.cell(row=row_idx, column=2, value=field.name)
                ws.cell(row=row_idx, column=3, value=field.caption or "")
                ws.cell(row=row_idx, column=4, value=field.data_type.value)
                ws.cell(row=row_idx, column=5, value=field.role.value)
                ws.cell(row=row_idx, column=6, value=field.default_aggregation.value)
                ws.cell(row=row_idx, column=7, value="Yes" if field.is_hidden else "No")
                row_idx += 1
        
        self._auto_width(ws)
    
    def _write_calculated_fields_sheet(self, ws, header_font, header_fill):
        """Write calculated fields information."""
        headers = ["Data Source", "Name", "Caption", "Formula", "Data Type", "Calculation Type", 
                   "Aggregations Used", "Functions Used", "Referenced Fields", "Complexity"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for ds in self.metadata.datasources:
            for calc in ds.calculated_fields:
                ws.cell(row=row_idx, column=1, value=ds.display_name)
                ws.cell(row=row_idx, column=2, value=calc.name)
                ws.cell(row=row_idx, column=3, value=calc.caption or "")
                ws.cell(row=row_idx, column=4, value=calc.formula[:500])  # Truncate long formulas
                ws.cell(row=row_idx, column=5, value=calc.data_type.value)
                ws.cell(row=row_idx, column=6, value=calc.calculation_type.value)
                ws.cell(row=row_idx, column=7, value=", ".join(calc.aggregations_used))
                ws.cell(row=row_idx, column=8, value=", ".join(calc.functions_used))
                ws.cell(row=row_idx, column=9, value=", ".join(calc.referenced_fields[:5]))
                ws.cell(row=row_idx, column=10, value=calc.complexity_score)
                row_idx += 1
        
        self._auto_width(ws, max_width=50)
    
    def _write_worksheets_sheet(self, ws, header_font, header_fill):
        """Write worksheets information."""
        headers = ["Sheet Name", "Title", "Data Source", "Chart Type", "Chart Type (Inferred)",
                   "Dimensions", "Measures", "# Filters", "Used in Dashboards"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for sheet in self.metadata.sheets:
            ws.cell(row=row_idx, column=1, value=sheet.name)
            ws.cell(row=row_idx, column=2, value=sheet.title or "")
            ws.cell(row=row_idx, column=3, value=sheet.datasource_name or "")
            
            if sheet.visual:
                ws.cell(row=row_idx, column=4, value=sheet.visual.chart_type.value)
                ws.cell(row=row_idx, column=5, value=sheet.visual.chart_type_inferred or "")
            
            ws.cell(row=row_idx, column=6, value=", ".join(sheet.dimensions_used[:5]))
            ws.cell(row=row_idx, column=7, value=", ".join(sheet.measures_used[:5]))
            ws.cell(row=row_idx, column=8, value=len(sheet.filters))
            ws.cell(row=row_idx, column=9, value=", ".join(sheet.used_in_dashboards))
            row_idx += 1
        
        self._auto_width(ws)
    
    def _write_filters_sheet(self, ws, header_font, header_fill):
        """Write filters information."""
        headers = ["Sheet", "Field", "Filter Type", "Is Context Filter", 
                   "Include Values", "Exclude Values", "Range Min", "Range Max",
                   "Condition", "Calculation Explanation"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for sheet in self.metadata.sheets:
            for filter in sheet.filters:
                ws.cell(row=row_idx, column=1, value=sheet.name)
                ws.cell(row=row_idx, column=2, value=filter.field)
                ws.cell(row=row_idx, column=3, value=filter.filter_type.value)
                ws.cell(row=row_idx, column=4, value="Yes" if filter.is_context_filter else "No")
                ws.cell(row=row_idx, column=5, value=", ".join(str(v) for v in filter.include_values[:5]))
                ws.cell(row=row_idx, column=6, value=", ".join(str(v) for v in filter.exclude_values[:5]))
                ws.cell(row=row_idx, column=7, value=str(filter.range_min) if filter.range_min else "")
                ws.cell(row=row_idx, column=8, value=str(filter.range_max) if filter.range_max else "")
                ws.cell(row=row_idx, column=9, value=filter.condition_formula or filter.formula or "")
                ws.cell(row=row_idx, column=10, value=filter.calculation_explanation or "")
                row_idx += 1
        
        self._auto_width(ws, max_width=50)
    
    def _write_dashboards_sheet(self, ws, header_font, header_fill):
        """Write dashboards information."""
        headers = ["Dashboard Name", "Title", "Width", "Height", "Layout Type",
                   "Worksheets", "# Zones", "# Actions", "Exposed Filters"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for dash in self.metadata.dashboards:
            ws.cell(row=row_idx, column=1, value=dash.name)
            ws.cell(row=row_idx, column=2, value=dash.title or "")
            ws.cell(row=row_idx, column=3, value=dash.width)
            ws.cell(row=row_idx, column=4, value=dash.height)
            ws.cell(row=row_idx, column=5, value=dash.layout_type)
            ws.cell(row=row_idx, column=6, value=", ".join(dash.worksheets))
            ws.cell(row=row_idx, column=7, value=len(dash.zones))
            ws.cell(row=row_idx, column=8, value=len(dash.actions))
            ws.cell(row=row_idx, column=9, value=", ".join(dash.exposed_filters))
            row_idx += 1
        
        self._auto_width(ws)
    
    def _write_parameters_sheet(self, ws, header_font, header_fill):
        """Write parameters information."""
        headers = ["Name", "Caption", "Data Type", "Current Value", 
                   "Allowable Type", "Allowable Values", "Min", "Max", "Step"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for param in self.metadata.parameters:
            ws.cell(row=row_idx, column=1, value=param.name)
            ws.cell(row=row_idx, column=2, value=param.caption or "")
            ws.cell(row=row_idx, column=3, value=param.data_type.value)
            ws.cell(row=row_idx, column=4, value=str(param.current_value) if param.current_value else "")
            ws.cell(row=row_idx, column=5, value=param.allowable_values_type)
            ws.cell(row=row_idx, column=6, value=", ".join(str(v) for v in param.allowable_values[:5]))
            ws.cell(row=row_idx, column=7, value=str(param.range_min) if param.range_min else "")
            ws.cell(row=row_idx, column=8, value=str(param.range_max) if param.range_max else "")
            ws.cell(row=row_idx, column=9, value=str(param.step_size) if param.step_size else "")
            row_idx += 1
        
        self._auto_width(ws)
    
    def _write_relationships_sheet(self, ws, header_font, header_fill):
        """Write relationships information."""
        headers = ["Relationship Type", "Source Type", "Source Name", 
                   "Target Type", "Target Name", "Description"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for rel in self.metadata.relationships:
            ws.cell(row=row_idx, column=1, value=rel.relationship_type)
            ws.cell(row=row_idx, column=2, value=rel.source_type)
            ws.cell(row=row_idx, column=3, value=rel.source_name)
            ws.cell(row=row_idx, column=4, value=rel.target_type)
            ws.cell(row=row_idx, column=5, value=rel.target_name)
            ws.cell(row=row_idx, column=6, value=rel.description or "")
            row_idx += 1
        
        self._auto_width(ws, max_width=60)
    
    def _write_metrics_sheet(self, ws, header_font, header_fill):
        """
        Write flattened metrics information - one row per metric-worksheet combination.
        
        This sheet provides a denormalized view where each metric usage in a worksheet
        is a unique row with all context (calculation, filters, dashboard, etc.).
        """
        headers = [
            "Metric Name", "Metric Caption", "Metric Type", 
            "Data Source", "Worksheet", "Chart Type", "Shelf Position",
            "Formula", "Formula (Readable)", "Calculation Type",
            "Data Type", "Aggregation Used", "Aggregations in Formula",
            "Functions Used", "Referenced Fields", "Referenced Parameters",
            "LOD Type", "LOD Dimensions", "LOD Expression",
            "Filters Applied", "Filter Details (Summary)",
            "Dashboards", "Complexity Score"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        for metric in self.metadata.metric_rows:
            ws.cell(row=row_idx, column=1, value=metric.metric_name)
            ws.cell(row=row_idx, column=2, value=metric.metric_caption or "")
            ws.cell(row=row_idx, column=3, value=metric.metric_type)
            ws.cell(row=row_idx, column=4, value=metric.datasource_caption or metric.datasource_name or "")
            ws.cell(row=row_idx, column=5, value=metric.worksheet_name)
            ws.cell(row=row_idx, column=6, value=metric.chart_type or "")
            ws.cell(row=row_idx, column=7, value=metric.shelf_position or "")
            ws.cell(row=row_idx, column=8, value=(metric.formula[:500] if metric.formula else ""))
            ws.cell(row=row_idx, column=9, value=(metric.formula_readable[:300] if metric.formula_readable else ""))
            ws.cell(row=row_idx, column=10, value=metric.calculation_type or "")
            ws.cell(row=row_idx, column=11, value=metric.data_type or "")
            ws.cell(row=row_idx, column=12, value=metric.aggregation_used or "")
            ws.cell(row=row_idx, column=13, value=", ".join(metric.aggregations_in_formula) if metric.aggregations_in_formula else "")
            ws.cell(row=row_idx, column=14, value=", ".join(metric.functions_used) if metric.functions_used else "")
            ws.cell(row=row_idx, column=15, value=", ".join(metric.referenced_fields[:5]) if metric.referenced_fields else "")
            ws.cell(row=row_idx, column=16, value=", ".join(metric.referenced_parameters) if metric.referenced_parameters else "")
            ws.cell(row=row_idx, column=17, value=metric.lod_type or "")
            ws.cell(row=row_idx, column=18, value=", ".join(metric.lod_dimensions) if metric.lod_dimensions else "")
            ws.cell(row=row_idx, column=19, value=metric.lod_expression or "")
            ws.cell(row=row_idx, column=20, value=", ".join(metric.filters_applied) if metric.filters_applied else "")
            
            # Summarize filter details
            filter_summary = ""
            if metric.filter_details:
                summaries = []
                for f in metric.filter_details[:3]:  # Limit to first 3 filters
                    summary = f"{f.get('field', '')}: {f.get('explanation', f.get('type', ''))}"
                    summaries.append(summary)
                filter_summary = "; ".join(summaries)
                if len(metric.filter_details) > 3:
                    filter_summary += f" (+{len(metric.filter_details) - 3} more)"
            ws.cell(row=row_idx, column=21, value=filter_summary)
            
            ws.cell(row=row_idx, column=22, value=", ".join(metric.dashboards_containing_worksheet) if metric.dashboards_containing_worksheet else "")
            ws.cell(row=row_idx, column=23, value=metric.complexity_score)
            row_idx += 1
        
        self._auto_width(ws, max_width=50)
    
    def _write_kpi_summary_sheet(self, ws, header_font, header_fill):
        """
        Write a business-friendly KPI/Metric summary for leadership review.
        
        This sheet provides a clear overview of all KPIs and calculated metrics
        with their exact calculation logic, usage across worksheets, and context.
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Business-friendly headers
        headers = [
            "KPI / Metric Name",
            "Display Name", 
            "Metric Type",
            "Calculation Logic (Formula)",
            "Formula Description",
            "Calculation Category",
            "Data Type",
            "Aggregations Used",
            "Functions Used",
            "Dependent Fields",
            "Parameters Referenced",
            "Used in Worksheets",
            "Worksheet Count",
            "Used in Dashboards",
            "Filters Applied (Summary)",
            "LOD Expression Type",
            "Complexity Level",
            "Complexity Score"
        ]
        
        # Style definitions
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Build a comprehensive KPI summary from all sources
        kpi_data = {}  # key: metric_name, value: aggregated info
        
        # First, collect all calculated fields from datasources
        for ds in self.metadata.datasources:
            for calc in ds.calculated_fields:
                key = calc.name
                if key not in kpi_data:
                    kpi_data[key] = {
                        "name": calc.name,
                        "caption": calc.caption,
                        "metric_type": "Calculated Field",
                        "formula": calc.formula,
                        "formula_readable": calc.formula_readable or calc.formula,
                        "calculation_type": calc.calculation_type.value if calc.calculation_type else "simple",
                        "data_type": calc.data_type.value if calc.data_type else "unknown",
                        "aggregations": calc.aggregations_used,
                        "functions": calc.functions_used,
                        "referenced_fields": calc.referenced_fields,
                        "referenced_parameters": calc.referenced_parameters,
                        "worksheets": set(),
                        "dashboards": set(),
                        "filters": set(),
                        "lod_type": calc.lod_type,
                        "complexity_score": calc.complexity_score,
                        "datasource": ds.caption or ds.name,
                    }
        
        # Add measure fields
        for ds in self.metadata.datasources:
            for field in ds.fields:
                if field.role.value == "measure":
                    key = field.name
                    if key not in kpi_data:
                        kpi_data[key] = {
                            "name": field.name,
                            "caption": field.caption,
                            "metric_type": "Measure",
                            "formula": None,
                            "formula_readable": f"[{field.name}]" + (f" (Default: {field.default_aggregation.value})" if field.default_aggregation else ""),
                            "calculation_type": "base_measure",
                            "data_type": field.data_type.value if field.data_type else "unknown",
                            "aggregations": [field.default_aggregation.value] if field.default_aggregation and field.default_aggregation.value != "none" else [],
                            "functions": [],
                            "referenced_fields": [],
                            "referenced_parameters": [],
                            "worksheets": set(),
                            "dashboards": set(),
                            "filters": set(),
                            "lod_type": None,
                            "complexity_score": 0,
                            "datasource": ds.caption or ds.name,
                        }
        
        # Enrich with worksheet and dashboard usage from metric_rows
        for metric in self.metadata.metric_rows:
            key = metric.metric_name
            if key in kpi_data:
                kpi_data[key]["worksheets"].add(metric.worksheet_name)
                for dash in metric.dashboards_containing_worksheet:
                    kpi_data[key]["dashboards"].add(dash)
                for f in metric.filters_applied:
                    kpi_data[key]["filters"].add(f)
        
        # Also check sheets directly for usage
        for sheet in self.metadata.sheets:
            for field_name in sheet.all_fields_used:
                if field_name in kpi_data:
                    kpi_data[field_name]["worksheets"].add(sheet.name)
            for field_name in sheet.measures_used:
                if field_name in kpi_data:
                    kpi_data[field_name]["worksheets"].add(sheet.name)
        
        # Map complexity score to business-friendly levels
        def get_complexity_level(score):
            if score >= 80:
                return "Very Complex"
            elif score >= 60:
                return "Complex"
            elif score >= 40:
                return "Moderate"
            elif score >= 20:
                return "Simple"
            else:
                return "Basic"
        
        # Generate formula description
        def get_formula_description(calc_type, lod_type, aggregations, functions):
            parts = []
            if calc_type == "lod_fixed":
                parts.append("Level of Detail (FIXED) calculation")
            elif calc_type == "lod_include":
                parts.append("Level of Detail (INCLUDE) calculation")
            elif calc_type == "lod_exclude":
                parts.append("Level of Detail (EXCLUDE) calculation")
            elif calc_type == "table_calc":
                parts.append("Table calculation (computed in viz)")
            elif calc_type == "aggregate":
                parts.append("Aggregate calculation")
            elif calc_type == "base_measure":
                parts.append("Base measure field")
            else:
                parts.append("Row-level calculation")
            
            if aggregations:
                parts.append(f"Uses: {', '.join(aggregations)}")
            if "IF" in functions or "IIF" in functions:
                parts.append("Contains conditional logic")
            if "CASE" in functions:
                parts.append("Contains CASE statement")
            
            return "; ".join(parts) if parts else "Standard calculation"
        
        # Sort KPIs: calculated fields first (by complexity), then measures
        sorted_kpis = sorted(
            kpi_data.values(),
            key=lambda x: (
                0 if x["metric_type"] == "Calculated Field" else 1,
                -x["complexity_score"],
                x["name"]
            )
        )
        
        # Write data rows
        row_idx = 2
        for kpi in sorted_kpis:
            worksheets_list = sorted(kpi["worksheets"]) if kpi["worksheets"] else []
            dashboards_list = sorted(kpi["dashboards"]) if kpi["dashboards"] else []
            filters_list = sorted(kpi["filters"]) if kpi["filters"] else []
            
            formula_desc = get_formula_description(
                kpi["calculation_type"],
                kpi["lod_type"],
                kpi["aggregations"],
                kpi["functions"]
            )
            
            ws.cell(row=row_idx, column=1, value=kpi["name"])
            ws.cell(row=row_idx, column=2, value=kpi["caption"] or kpi["name"])
            ws.cell(row=row_idx, column=3, value=kpi["metric_type"])
            
            # Formula - truncate for readability but include full formula
            formula = kpi["formula_readable"] or kpi["formula"] or ""
            ws.cell(row=row_idx, column=4, value=formula[:1000] if formula else "N/A")
            
            ws.cell(row=row_idx, column=5, value=formula_desc)
            ws.cell(row=row_idx, column=6, value=kpi["calculation_type"].replace("_", " ").title())
            ws.cell(row=row_idx, column=7, value=kpi["data_type"])
            ws.cell(row=row_idx, column=8, value=", ".join(kpi["aggregations"]) if kpi["aggregations"] else "None")
            ws.cell(row=row_idx, column=9, value=", ".join(kpi["functions"][:10]) if kpi["functions"] else "None")
            ws.cell(row=row_idx, column=10, value=", ".join(kpi["referenced_fields"][:10]) if kpi["referenced_fields"] else "None")
            ws.cell(row=row_idx, column=11, value=", ".join(kpi["referenced_parameters"]) if kpi["referenced_parameters"] else "None")
            ws.cell(row=row_idx, column=12, value=", ".join(worksheets_list[:5]) + (f" (+{len(worksheets_list)-5} more)" if len(worksheets_list) > 5 else "") if worksheets_list else "Not used")
            ws.cell(row=row_idx, column=13, value=len(worksheets_list))
            ws.cell(row=row_idx, column=14, value=", ".join(dashboards_list) if dashboards_list else "None")
            ws.cell(row=row_idx, column=15, value=", ".join(filters_list[:3]) + (f" (+{len(filters_list)-3} more)" if len(filters_list) > 3 else "") if filters_list else "None")
            ws.cell(row=row_idx, column=16, value=kpi["lod_type"] or "N/A")
            ws.cell(row=row_idx, column=17, value=get_complexity_level(kpi["complexity_score"]))
            ws.cell(row=row_idx, column=18, value=kpi["complexity_score"])
            
            row_idx += 1
        
        # Set column widths for readability
        column_widths = {
            'A': 35,  # KPI Name
            'B': 30,  # Display Name
            'C': 18,  # Metric Type
            'D': 80,  # Formula
            'E': 50,  # Formula Description
            'F': 18,  # Calculation Category
            'G': 12,  # Data Type
            'H': 25,  # Aggregations
            'I': 35,  # Functions
            'J': 40,  # Dependent Fields
            'K': 25,  # Parameters
            'L': 50,  # Worksheets
            'M': 12,  # Worksheet Count
            'N': 40,  # Dashboards
            'O': 40,  # Filters
            'P': 15,  # LOD Type
            'Q': 15,  # Complexity Level
            'R': 12,  # Complexity Score
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
    
    def _auto_width(self, ws, max_width: int = 40):
        """Auto-adjust column widths."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column].width = adjusted_width
    
    def to_html(self, output_path: str, include_details: bool = True):
        """
        Export metadata to interactive HTML report.
        
        Args:
            output_path: Path to save the HTML file
            include_details: Include detailed sections
        """
        html_content = self._generate_html(include_details)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _generate_html(self, include_details: bool = True) -> str:
        """Generate HTML content."""
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tableau Metadata Report - {self.metadata.name}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            line-height: 1.6;
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }}
        h1 {{ color: #1f3a60; border-bottom: 3px solid #e97627; padding-bottom: 10px; }}
        h2 {{ color: #1f3a60; margin-top: 30px; }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
        }}
        .stat-box {{
            background: #e97627;
            color: white;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-value {{ font-size: 2em; font-weight: bold; }}
        .stat-label {{ font-size: 0.9em; opacity: 0.9; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}
        th, td {{
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background: #1f3a60;
            color: white;
        }}
        tr:hover {{ background: #f5f5f5; }}
        .tag {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 0.85em;
            margin: 2px;
        }}
        .tag-dimension {{ background: #4CAF50; color: white; }}
        .tag-measure {{ background: #2196F3; color: white; }}
        .tag-calc {{ background: #9C27B0; color: white; }}
        .tag-lod {{ background: #FF5722; color: white; }}
        .formula {{ 
            font-family: monospace;
            background: #f0f0f0;
            padding: 5px 10px;
            border-radius: 4px;
            overflow-x: auto;
            display: block;
            max-width: 600px;
        }}
        .collapsible {{
            cursor: pointer;
            padding: 15px;
            background: #1f3a60;
            color: white;
            border: none;
            width: 100%;
            text-align: left;
            outline: none;
            font-size: 1.1em;
            border-radius: 4px;
            margin: 5px 0;
        }}
        .collapsible:after {{ content: '\\002B'; float: right; }}
        .collapsible.active:after {{ content: '\\2212'; }}
        .content {{
            padding: 0 18px;
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.2s ease-out;
            background: white;
        }}
    </style>
</head>
<body>
    <h1>📊 Tableau Metadata Report</h1>
    
    <div class="card">
        <h2>Workbook: {self.metadata.name}</h2>
        <p><strong>Version:</strong> {self.metadata.version or 'N/A'}</p>
        <p><strong>Source:</strong> {self.metadata.source_file or 'N/A'}</p>
        <p><strong>Extracted:</strong> {self.metadata.extraction_timestamp}</p>
    </div>
    
    <div class="card">
        <h2>📈 Summary Statistics</h2>
        <div class="stats-grid">
            <div class="stat-box">
                <div class="stat-value">{len(self.metadata.datasources)}</div>
                <div class="stat-label">Data Sources</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{self.metadata.total_sheets}</div>
                <div class="stat-label">Worksheets</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{self.metadata.total_dashboards}</div>
                <div class="stat-label">Dashboards</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{self.metadata.total_fields}</div>
                <div class="stat-label">Fields</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{self.metadata.total_calculated_fields}</div>
                <div class="stat-label">Calculated Fields</div>
            </div>
            <div class="stat-box">
                <div class="stat-value">{self.metadata.total_parameters}</div>
                <div class="stat-label">Parameters</div>
            </div>
        </div>
    </div>
    
    {self._generate_sheets_section() if include_details else ''}
    {self._generate_calcs_section() if include_details else ''}
    {self._generate_filters_section() if include_details else ''}
    {self._generate_relationships_section() if include_details else ''}
    
    <script>
        document.querySelectorAll('.collapsible').forEach(btn => {{
            btn.addEventListener('click', function() {{
                this.classList.toggle('active');
                const content = this.nextElementSibling;
                if (content.style.maxHeight) {{
                    content.style.maxHeight = null;
                }} else {{
                    content.style.maxHeight = content.scrollHeight + 'px';
                }}
            }});
        }});
    </script>
</body>
</html>
"""
    
    def _generate_sheets_section(self) -> str:
        """Generate sheets section HTML."""
        rows = ""
        for sheet in self.metadata.sheets:
            chart_type = sheet.visual.chart_type.value if sheet.visual else "N/A"
            rows += f"""
            <tr>
                <td>{sheet.name}</td>
                <td>{chart_type}</td>
                <td>{', '.join(sheet.dimensions_used[:3])}</td>
                <td>{', '.join(sheet.measures_used[:3])}</td>
                <td>{len(sheet.filters)}</td>
            </tr>
            """
        
        return f"""
        <div class="card">
            <h2>📋 Worksheets</h2>
            <table>
                <tr>
                    <th>Name</th>
                    <th>Chart Type</th>
                    <th>Dimensions</th>
                    <th>Measures</th>
                    <th>Filters</th>
                </tr>
                {rows}
            </table>
        </div>
        """
    
    def _generate_calcs_section(self) -> str:
        """Generate calculated fields section HTML."""
        rows = ""
        for ds in self.metadata.datasources:
            for calc in ds.calculated_fields:
                type_tag = "tag-lod" if calc.is_lod else "tag-calc"
                rows += f"""
                <tr>
                    <td>{calc.display_name}</td>
                    <td><span class="tag {type_tag}">{calc.calculation_type.value}</span></td>
                    <td><code class="formula">{calc.formula[:100]}{'...' if len(calc.formula) > 100 else ''}</code></td>
                    <td>{calc.complexity_score}</td>
                </tr>
                """
        
        return f"""
        <div class="card">
            <h2>🔢 Calculated Fields</h2>
            <table>
                <tr>
                    <th>Name</th>
                    <th>Type</th>
                    <th>Formula</th>
                    <th>Complexity</th>
                </tr>
                {rows}
            </table>
        </div>
        """
    
    def _generate_filters_section(self) -> str:
        """Generate filters section HTML."""
        rows = ""
        for sheet in self.metadata.sheets:
            for f in sheet.filters:
                rows += f"""
                <tr>
                    <td>{sheet.name}</td>
                    <td>{f.field}</td>
                    <td>{f.filter_type.value}</td>
                    <td>{f.calculation_explanation or 'N/A'}</td>
                </tr>
                """
        
        return f"""
        <div class="card">
            <h2>🔍 Filters</h2>
            <table>
                <tr>
                    <th>Sheet</th>
                    <th>Field</th>
                    <th>Type</th>
                    <th>Description</th>
                </tr>
                {rows}
            </table>
        </div>
        """
    
    def _generate_relationships_section(self) -> str:
        """Generate relationships section HTML."""
        rows = ""
        for rel in self.metadata.relationships[:50]:  # Limit to 50
            rows += f"""
            <tr>
                <td>{rel.relationship_type}</td>
                <td>{rel.source_type}: {rel.source_name}</td>
                <td>{rel.target_type}: {rel.target_name}</td>
            </tr>
            """
        
        return f"""
        <div class="card">
            <h2>🔗 Relationships</h2>
            <table>
                <tr>
                    <th>Type</th>
                    <th>Source</th>
                    <th>Target</th>
                </tr>
                {rows}
            </table>
            {f'<p><em>Showing first 50 of {len(self.metadata.relationships)} relationships</em></p>' if len(self.metadata.relationships) > 50 else ''}
        </div>
        """
    
    def to_summary(self) -> str:
        """
        Generate a text summary for console output.
        
        Returns:
            str: Summary text
        """
        lines = [
            "=" * 60,
            f"TABLEAU WORKBOOK METADATA: {self.metadata.name}",
            "=" * 60,
            "",
            "BASIC INFO",
            "-" * 40,
            f"Version: {self.metadata.version or 'N/A'}",
            f"Source: {self.metadata.source_file or 'N/A'}",
            f"Extraction Method: {self.metadata.extraction_method}",
            "",
            "STATISTICS",
            "-" * 40,
            f"Data Sources: {len(self.metadata.datasources)}",
            f"Worksheets: {self.metadata.total_sheets}",
            f"Dashboards: {self.metadata.total_dashboards}",
            f"Total Fields: {self.metadata.total_fields}",
            f"Calculated Fields: {self.metadata.total_calculated_fields}",
            f"Parameters: {self.metadata.total_parameters}",
            f"Total Filters: {self.metadata.total_filters}",
            "",
        ]
        
        # Sheets summary
        if self.metadata.sheets:
            lines.extend([
                "WORKSHEETS",
                "-" * 40,
            ])
            for sheet in self.metadata.sheets[:10]:
                chart_type = sheet.visual.chart_type.value if sheet.visual else "N/A"
                lines.append(f"  • {sheet.name} ({chart_type}) - {len(sheet.filters)} filters")
            if len(self.metadata.sheets) > 10:
                lines.append(f"  ... and {len(self.metadata.sheets) - 10} more")
            lines.append("")
        
        # Calculated fields summary
        calc_count = sum(len(ds.calculated_fields) for ds in self.metadata.datasources)
        if calc_count > 0:
            lines.extend([
                "CALCULATED FIELDS (Top 10 by complexity)",
                "-" * 40,
            ])
            all_calcs = []
            for ds in self.metadata.datasources:
                all_calcs.extend(ds.calculated_fields)
            
            sorted_calcs = sorted(all_calcs, key=lambda x: x.complexity_score, reverse=True)[:10]
            for calc in sorted_calcs:
                lines.append(f"  • {calc.display_name} [{calc.calculation_type.value}] (complexity: {calc.complexity_score})")
            lines.append("")
        
        # Dashboards summary
        if self.metadata.dashboards:
            lines.extend([
                "DASHBOARDS",
                "-" * 40,
            ])
            for dash in self.metadata.dashboards:
                lines.append(f"  • {dash.name} ({dash.width}x{dash.height}) - {len(dash.worksheets)} sheets, {len(dash.actions)} actions")
            lines.append("")
        
        lines.append("=" * 60)
        
        return "\n".join(lines)
